#!/usr/bin/env python3
"""
SkinVault by Chi — product image downloader
--------------------------------------------
Reads SkinVault_Image_Sourcing_Manifest.xlsx and downloads product images
into assets/products/ under the exact filenames the website expects.

For each manifest row it tries, in order:
  1. The "Official Image URL" column (direct download)
  2. The "Official Product Page" column — fetches the page and extracts
     the og:image / Shopify product image automatically
  3. Otherwise marks the row "manual" for the search-link workflow

Usage (from the website folder, next to the manifest):
    pip install requests openpyxl
    python download_product_images.py

Run it as many times as you like — files already downloaded are skipped,
so you can re-run after adding more URLs to the manifest.
"""
import os, re, sys, time

try:
    import requests
    from openpyxl import load_workbook
except ImportError:
    sys.exit("Please run:  pip install requests openpyxl   and try again.")

MANIFEST = "SkinVault_Image_Sourcing_Manifest.xlsx"
OUTDIR = os.path.join("assets", "products")
HEADERS = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
           "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"}

os.makedirs(OUTDIR, exist_ok=True)

def get_hyperlink(cell):
    return cell.hyperlink.target if cell and cell.hyperlink else None

def find_og_image(page_url):
    """Fetch a product page and pull the official product image URL."""
    try:
        r = requests.get(page_url, headers=HEADERS, timeout=25)
        r.raise_for_status()
    except Exception as e:
        return None, f"page fetch failed ({e})"
    html = r.text
    m = re.search(r'<meta[^>]+property=["\']og:image["\'][^>]+content=["\']([^"\']+)', html)
    if not m:
        m = re.search(r'<meta[^>]+content=["\']([^"\']+)["\'][^>]+property=["\']og:image["\']', html)
    if m:
        url = m.group(1)
        if url.startswith("//"): url = "https:" + url
        return url, None
    # Shopify fallback: first product image in JSON-LD or cdn.shopify link
    m = re.search(r'https://cdn\.shopify\.com/[^"\'\s\\]+\.(?:jpg|jpeg|png|webp)[^"\'\s\\]*', html)
    if m:
        return m.group(0), None
    return None, "no og:image found on page"

def download(url, path):
    try:
        r = requests.get(url, headers=HEADERS, timeout=40)
        r.raise_for_status()
        if len(r.content) < 5000:
            return "file too small — probably blocked"
        with open(path, "wb") as f:
            f.write(r.content)
        return None
    except Exception as e:
        return str(e)

def main():
    wb = load_workbook(MANIFEST)
    ws = wb.active
    ok = skipped = manual = failed = 0
    for row in ws.iter_rows(min_row=5):
        fn = row[4].value
        if not fn:
            continue
        dest = os.path.join(OUTDIR, fn)
        if os.path.exists(dest):
            skipped += 1
            continue
        img_url = get_hyperlink(row[9] if len(row) > 9 else None)   # col J: Official Image URL
        page_url = get_hyperlink(row[8] if len(row) > 8 else None)  # col I: Official Product Page
        err = None
        if not img_url and page_url:
            img_url, err = find_og_image(page_url)
            time.sleep(1)  # be polite to brand servers
        if not img_url:
            manual += 1
            print(f"  MANUAL   {fn}  ({err or 'no URL in manifest — use the search link'})")
            continue
        err = download(img_url, dest)
        if err:
            failed += 1
            print(f"  FAILED   {fn}  ({err})")
        else:
            ok += 1
            row[7].value = "Downloaded"           # col H: Status
            print(f"  SAVED    {fn}")
        time.sleep(1)
    wb.save(MANIFEST)
    print(f"\nDone. {ok} downloaded, {skipped} already present, "
          f"{manual} need the manual search-link workflow, {failed} failed.")
    print(f"Images are in {OUTDIR}/ — the website picks them up automatically.")

if __name__ == "__main__":
    main()
