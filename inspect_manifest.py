import openpyxl
from itertools import islice
wb = openpyxl.load_workbook('SkinVault_Image_Sourcing_Manifest.xlsx')
print('sheets=', wb.sheetnames)
ws = wb.active
print('max_row', ws.max_row, 'max_col', ws.max_column)
header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
print('header=', header)
for i, row in enumerate(islice(ws.iter_rows(min_row=2, max_row=min(ws.max_row, 22), values_only=True), 21), start=2):
    print(i, row)
