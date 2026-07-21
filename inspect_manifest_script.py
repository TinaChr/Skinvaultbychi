import os
from openpyxl import load_workbook

site_root = r"c:\Users\HP\Downloads\skinvault-website (1)"
manifest_path = os.path.join(site_root, "SkinVault_Image_Sourcing_Manifest.xlsx")
attached_dir = r"c:\Users\HP\Downloads\SkinvaultbyChi Product Catalogue -20260720T192435Z-1-001\SkinvaultbyChi Product Catalogue"

wb = load_workbook(manifest_path)
ws = wb.active
expected = [row[4].value.strip() for row in ws.iter_rows(min_row=5) if row[4].value]
attached = sorted(os.listdir(attached_dir))
exact_matches = sorted(set(expected) & set(attached))
missing = sorted(set(expected) - set(attached))

print(f"Expected filenames: {len(expected)}")
print(f"Attached files: {len(attached)}")
print(f"Exact filename matches: {len(exact_matches)}")
print(f"Missing from attached folder: {len(missing)}")

if exact_matches:
    print("\nExact matches:")
    for fn in exact_matches[:50]:
        print(fn)
    if len(exact_matches) > 50:
        print(f"...and {len(exact_matches) - 50} more exact matches")

print("\nMissing examples:")
for fn in missing[:50]:
    print(fn)
if len(missing) > 50:
    print(f"...and {len(missing) - 50} more missing filenames")

missing_file_path = os.path.join(site_root, 'missing_product_images.txt')
with open(missing_file_path, 'w', encoding='utf-8') as f:
    for fn in missing:
        f.write(fn + '\n')

print(f"\nSaved missing filename list to: {missing_file_path}")
